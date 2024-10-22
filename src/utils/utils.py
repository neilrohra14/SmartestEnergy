import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import subprocess

ARTIFACTS_DIR = os.path.join(os.getcwd(), 'artifacts')
EXCEL_FILE_PATH = os.path.join(ARTIFACTS_DIR, 'cleaned_data.xlsx')

def save_cleaned_data_to_excel(df, target_date, sheet_name=None, image_paths=None):

    if not os.path.exists(ARTIFACTS_DIR):
        os.makedirs(ARTIFACTS_DIR)

    if sheet_name is None:
        sheet_name = f"Report_{target_date}"

    if os.path.exists(EXCEL_FILE_PATH):
        with pd.ExcelWriter(EXCEL_FILE_PATH, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        if image_paths:
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet = workbook[sheet_name]

            row = 5
            for image_path in image_paths:
                img = Image(image_path)

                img.width = 500  
                img.height = 300 

                cell_position = f'B{row}'
                sheet.add_image(img, cell_position)
                row += 20 

            workbook.save(EXCEL_FILE_PATH)
    else:
        
        with pd.ExcelWriter(EXCEL_FILE_PATH, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        if image_paths:
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet = workbook[sheet_name]

            row = 5
            for image_path in image_paths:
                img = Image(image_path)

                img.width = 500
                img.height = 300

                cell_position = f'B{row}'
                sheet.add_image(img, cell_position)
                row += 20  #row spacing

            workbook.save(EXCEL_FILE_PATH)

    print(f"Data for {target_date} has been saved to {EXCEL_FILE_PATH} in sheet {sheet_name}")

def open_excel_file(file_path):
    try:
        if os.name == 'posix':  # macOS/Linux
            subprocess.run(['open', file_path], check=True)
        elif os.name == 'nt':  # Windows
            os.startfile(file_path)
    except Exception as e:
        print(f"Failed to open the Excel file: {str(e)}")
